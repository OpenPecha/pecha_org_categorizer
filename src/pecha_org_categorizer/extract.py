from pathlib import Path
from typing import List, Union

from openpyxl import load_workbook


class Extractor:
    def __init__(self):
        self.extracted_info = []

    @staticmethod
    def extract_xlsx(input_xlsx: Path):
        wb = load_workbook(input_xlsx)
        ws = wb.active
        extracted_info = []
        for row in ws.iter_rows(values_only=True):
            extracted_info.append(row)
        return extracted_info

    def extract(self, input_xlsx: Path):
        """
        Read the xlsx file and extract the information
        """
        self.extracted_info = self.extract_xlsx(input_xlsx)
        all_category = []
        curr_category: List[Union[str, None]] = []

        for row_idx, row in enumerate(self.extracted_info):
            # Find the first non-None value and its index
            for col_idx, data in enumerate(row):
                if data is not None:
                    break
            else:
                continue  # Skip rows with all None values

            curr_category_len = len(curr_category)

            if curr_category_len < col_idx:
                raise ValueError(
                    f"Data is not in the right format. Please check line number {row_idx} in the xlsx file."
                )

            # Update or extend the current category list
            if curr_category_len == col_idx:
                curr_category.append(data)
            else:
                curr_category[col_idx] = data

            # Set the trailing elements to None if needed
            curr_category[col_idx + 1 :] = [None] * (  # noqa
                curr_category_len - col_idx - 1
            )

            # Add non-empty elements to the result
            non_empty_curr_category = [x for x in curr_category if x is not None]
            all_category.append(non_empty_curr_category)

        return all_category
